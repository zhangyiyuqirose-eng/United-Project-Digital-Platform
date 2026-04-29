"""Excel report generation service using openpyxl."""

from __future__ import annotations

import io
import uuid
from datetime import date, datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.business.models import Contract, ContractPayment
from app.models.project.models import (
    Project,
    ProjectTask,
    ProjectMilestone,
    ProjectRisk,
)
from app.models.resource.models import PoolMembership, ResourcePool
from app.models.timesheet.models import Timesheet


def _style_header_cell(cell, font_color="FFFFFF"):
    font = Font(bold=True, color=font_color, size=11)
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.font = font
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def _style_data_cell(cell, is_number=False):
    cell.alignment = Alignment(horizontal="center" if is_number else "left", vertical="center")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def _safe_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def _safe_float(value) -> float:
    if value is None:
        return 0.0
    return float(value)


class ReportService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    async def generate_project_report(self, project_id: str | None = None) -> bytes:
        projects = await self._fetch_projects(project_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "Project Report"

        headers = [
            "Project Name", "Project Code", "Status", "Manager",
            "Department", "Start Date", "End Date", "Budget",
            "Progress %", "CPI", "SPI", "Tasks", "Milestones",
            "Risks", "Customer",
        ]
        ws.append(headers)
        for col_idx, h in enumerate(headers, 1):
            _style_header_cell(ws.cell(row=1, column=col_idx))

        for p in projects:
            task_count = await self._count_tasks(p["project_id"])
            milestone_count = await self._count_milestones(p["project_id"])
            risk_count = await self._count_risks(p["project_id"])
            row = [
                p.get("project_name", ""),
                p.get("project_code", ""),
                p.get("status", ""),
                p.get("manager_name", ""),
                p.get("department_name", ""),
                _safe_str(p.get("start_date")),
                _safe_str(p.get("end_date")),
                _safe_float(p.get("budget")),
                p.get("progress", 0),
                _safe_float(p.get("evm_cpi")),
                _safe_float(p.get("evm_spi")),
                task_count,
                milestone_count,
                risk_count,
                p.get("customer", ""),
            ]
            ws.append(row)
            for col_idx in range(1, len(headers) + 1):
                _style_data_cell(ws.cell(row=ws.max_row, column=col_idx))

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["N"].width = 12

        return self._save_workbook(wb)

    async def generate_cost_report(self, project_id: str | None = None) -> bytes:
        projects = await self._fetch_projects(project_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "Cost Report"

        headers = [
            "Project Name", "Project Code", "Budget",
            "Contract Count", "Payment Count", "Paid Amount",
            "Pending Amount", "Budget Variance", "Variance Rate %",
        ]
        ws.append(headers)
        for col_idx, h in enumerate(headers, 1):
            _style_header_cell(ws.cell(row=1, column=col_idx))

        for p in projects:
            budget = _safe_float(p.get("budget"))
            contracts = await self._count_contracts(p["project_id"])
            payments = await self._fetch_payments(p["project_id"])
            paid = sum(_safe_float(pay.get("amount")) for pay in payments)
            pending = sum(
                _safe_float(pay.get("amount"))
                for pay in payments
                if pay.get("status") != "paid"
            )
            variance = budget - paid
            variance_rate = (variance / budget * 100) if budget > 0 else 0.0
            row = [
                p.get("project_name", ""),
                p.get("project_code", ""),
                budget,
                contracts,
                len(payments),
                round(paid, 2),
                round(pending, 2),
                round(variance, 2),
                round(variance_rate, 2),
            ]
            ws.append(row)
            for col_idx in range(1, len(headers) + 1):
                _style_data_cell(ws.cell(row=ws.max_row, column=col_idx), is_number=True)

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["F"].width = 16
        ws.column_dimensions["H"].width = 18

        return self._save_workbook(wb)

    async def generate_timesheet_report(
        self, staff_id: str | None = None, month: str | None = None
    ) -> bytes:
        timesheets = await self._fetch_timesheets(staff_id, month)
        wb = Workbook()
        ws = wb.active
        ws.title = "Timesheet Report"

        headers = [
            "Staff ID", "Staff Name", "Project", "Work Month",
            "Total Hours", "Approved Hours", "Rejected Hours", "Status",
        ]
        ws.append(headers)
        for col_idx, h in enumerate(headers, 1):
            _style_header_cell(ws.cell(row=1, column=col_idx))

        for ts in timesheets:
            row = [
                ts.get("staff_id", ""),
                ts.get("staff_name", ""),
                ts.get("project_name", ""),
                ts.get("work_month", ""),
                _safe_float(ts.get("total_hours")),
                _safe_float(ts.get("approved_hours")),
                _safe_float(ts.get("rejected_hours")),
                ts.get("status", ""),
            ]
            ws.append(row)
            for col_idx in range(1, len(headers) + 1):
                _style_data_cell(ws.cell(row=ws.max_row, column=col_idx))

        return self._save_workbook(wb)

    async def generate_resource_report(self, pool_id: str | None = None) -> bytes:
        pools = await self._fetch_pools(pool_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "Resource Report"

        headers = [
            "Pool Name", "Total Staff", "Active Staff",
            "Assigned Staff", "Utilization Rate %",
        ]
        ws.append(headers)
        for col_idx, h in enumerate(headers, 1):
            _style_header_cell(ws.cell(row=1, column=col_idx))

        for pool in pools:
            total = pool.get("total_staff", 0)
            assigned = pool.get("assigned_staff", 0)
            utilization = round((assigned / total * 100) if total > 0 else 0, 1)
            row = [
                pool.get("pool_name", ""),
                total,
                pool.get("active_staff", 0),
                assigned,
                utilization,
            ]
            ws.append(row)
            for col_idx in range(1, len(headers) + 1):
                _style_data_cell(ws.cell(row=ws.max_row, column=col_idx), is_number=True)

        return self._save_workbook(wb)

    async def generate_portfolio_report(self) -> bytes:
        projects = await self._fetch_projects(None)
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Project Summary"
        headers = ["Project Name", "Code", "Status", "Budget", "Progress %", "CPI", "SPI"]
        ws1.append(headers)
        for col_idx, h in enumerate(headers, 1):
            _style_header_cell(ws1.cell(row=1, column=col_idx))

        for p in projects:
            row = [
                p.get("project_name", ""),
                p.get("project_code", ""),
                p.get("status", ""),
                _safe_float(p.get("budget")),
                p.get("progress", 0),
                _safe_float(p.get("evm_cpi")),
                _safe_float(p.get("evm_spi")),
            ]
            ws1.append(row)
            for col_idx in range(1, len(headers) + 1):
                _style_data_cell(ws1.cell(row=ws1.max_row, column=col_idx))

        total_budget = sum(_safe_float(p.get("budget")) for p in projects)
        avg_progress = (
            sum(p.get("progress", 0) for p in projects) / len(projects) if projects else 0
        )
        ws2 = wb.create_sheet("Portfolio Summary")
        summary_headers = ["Metric", "Value"]
        ws2.append(summary_headers)
        for col_idx, h in enumerate(summary_headers, 1):
            _style_header_cell(ws2.cell(row=1, column=col_idx))

        summary = [
            ("Total Projects", len(projects)),
            ("Total Budget", round(total_budget, 2)),
            ("Avg Progress %", round(avg_progress, 1)),
        ]
        for label, value in summary:
            ws2.append([label, value])
            _style_data_cell(ws2.cell(row=ws2.max_row, column=1))
            _style_data_cell(ws2.cell(row=ws2.max_row, column=2), is_number=True)

        return self._save_workbook(wb)

    # ---- Data fetching helpers ----

    async def _fetch_projects(self, project_id: str | None) -> list[dict]:
        stmt = select(Project).order_by(Project.create_time.desc())
        if project_id:
            stmt = stmt.where(Project.project_id == project_id)
        result = await self.db.execute(stmt)
        projects = result.scalars().all()
        return [
            {
                "project_id": p.project_id,
                "project_name": p.project_name,
                "project_code": p.project_code,
                "status": p.status,
                "manager_name": p.manager_name,
                "department_name": p.department_name,
                "start_date": p.start_date,
                "end_date": p.end_date,
                "budget": p.budget,
                "progress": p.progress,
                "evm_cpi": p.evm_cpi,
                "evm_spi": p.evm_spi,
                "customer": p.customer,
            }
            for p in projects
        ]

    async def _count_tasks(self, project_id: str) -> int:
        stmt = select(func.count()).select_from(ProjectTask).where(
            ProjectTask.project_id == project_id
        )
        return (await self.db.execute(stmt)).scalar() or 0

    async def _count_milestones(self, project_id: str) -> int:
        stmt = select(func.count()).select_from(ProjectMilestone).where(
            ProjectMilestone.project_id == project_id
        )
        return (await self.db.execute(stmt)).scalar() or 0

    async def _count_risks(self, project_id: str) -> int:
        stmt = select(func.count()).select_from(ProjectRisk).where(
            ProjectRisk.project_id == project_id
        )
        return (await self.db.execute(stmt)).scalar() or 0

    async def _count_contracts(self, project_id: str) -> int:
        stmt = select(func.count()).select_from(Contract).where(
            Contract.project_id == project_id
        )
        return (await self.db.execute(stmt)).scalar() or 0

    async def _fetch_payments(self, project_id: str) -> list[dict]:
        stmt = (
            select(ContractPayment)
            .join(Contract, ContractPayment.contract_id == Contract.contract_id)
            .where(Contract.project_id == project_id)
        )
        result = await self.db.execute(stmt)
        payments = result.scalars().all()
        return [
            {
                "amount": p.amount,
                "status": p.status,
            }
            for p in payments
        ]

    async def _fetch_timesheets(
        self, staff_id: str | None, month: str | None
    ) -> list[dict]:
        stmt = select(Timesheet).order_by(Timesheet.create_time.desc())
        if staff_id:
            stmt = stmt.where(Timesheet.staff_id == staff_id)
        if month:
            stmt = stmt.where(Timesheet.work_month == month)
        result = await self.db.execute(stmt)
        timesheets = result.scalars().all()
        return [
            {
                "staff_id": ts.staff_id,
                "staff_name": ts.staff_name,
                "project_name": ts.project_name,
                "work_month": ts.work_month,
                "total_hours": ts.total_hours,
                "approved_hours": ts.approved_hours,
                "rejected_hours": ts.rejected_hours,
                "status": ts.status,
            }
            for ts in timesheets
        ]

    async def _fetch_pools(self, pool_id: str | None) -> list[dict]:
        stmt = (
            select(
                ResourcePool.pool_id,
                ResourcePool.pool_name,
                func.count(PoolMembership.person_id).label("total_staff"),
            )
            .outerjoin(PoolMembership, ResourcePool.pool_id == PoolMembership.pool_id)
            .group_by(ResourcePool.pool_id, ResourcePool.pool_name)
        )
        if pool_id:
            stmt = stmt.where(ResourcePool.pool_id == pool_id)
        result = await self.db.execute(stmt)
        rows = result.all()

        pool_data = []
        for row in rows:
            assigned_stmt = (
                select(func.count())
                .select_from(PoolMembership)
                .where(
                    PoolMembership.pool_id == row.pool_id,
                    PoolMembership.status == "assigned",
                )
            )
            active_stmt = (
                select(func.count())
                .select_from(PoolMembership)
                .where(
                    PoolMembership.pool_id == row.pool_id,
                    PoolMembership.status.in_(["assigned", "available"]),
                )
            )
            assigned = (await self.db.execute(assigned_stmt)).scalar() or 0
            active = (await self.db.execute(active_stmt)).scalar() or 0
            pool_data.append(
                {
                    "pool_name": row.pool_name,
                    "total_staff": row.total_staff,
                    "active_staff": active,
                    "assigned_staff": assigned,
                }
            )
        return pool_data

    @staticmethod
    def _save_workbook(wb: Workbook) -> bytes:
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
